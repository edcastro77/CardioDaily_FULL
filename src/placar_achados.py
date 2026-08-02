"""
placar_achados.py — o veredito sobre QUAL MODELO ENXERGA (02/Ago/2026).

Lê a planilha marcada pelo Dr. Eduardo (outputs/PROVA_PROMPTS/achados_para_marcar.xlsx|csv) e
responde a pergunta que a contagem de tabelas nunca respondeu:

    quem produz mais ACHADO DE PERÍCIA VERDADEIRO — a contradição interna, o número que não fecha,
    o método ausente, a extrapolação — e quem INVENTA achado que não existe?

Duas medidas, e a segunda importa mais que a primeira:
  • ACHADOS REAIS por documento  → o que o assinante ganha
  • TAXA DE INVENÇÃO             → quantos dos "achados" do modelo são falsos.
    Um modelo que inventa contradição é PIOR que um que não acha nada: ele destrói a confiança
    no produto inteiro, que é a única coisa que o CardioDaily tem para vender.

Uso: python src/placar_achados.py
"""
import os
import csv
import sys
from collections import defaultdict

_HERE = os.path.dirname(os.path.abspath(__file__))
SAIDA = os.path.join(os.path.dirname(_HERE), "outputs", "PROVA_PROMPTS")

SIM = {"s", "sim", "y", "yes", "1", "v", "x"}
NAO = {"n", "nao", "não", "0", "f"}
PESO = {"alto": 3, "medio": 2, "médio": 2, "baixo": 1, "": 1}


def carregar():
    xlsx = os.path.join(SAIDA, "achados_para_marcar.xlsx")
    csv_p = os.path.join(SAIDA, "achados_para_marcar.csv")
    if os.path.exists(xlsx):
        try:
            from openpyxl import load_workbook
            ws = load_workbook(xlsx, data_only=True).active
            cab = [c.value for c in ws[1]]
            return [dict(zip(cab, r)) for r in ws.iter_rows(min_row=2, values_only=True) if r[0]], "xlsx"
        except ImportError:
            pass
    if os.path.exists(csv_p):
        return list(csv.DictReader(open(csv_p, encoding="utf-8-sig"))), "csv"
    return [], None


def _v(x):
    return str(x or "").strip().lower()


def main():
    linhas, fonte = carregar()
    if not linhas:
        print("Não achei a planilha. Rode antes: python src/achados.py"); return 1

    marcadas = [l for l in linhas if _v(l.get("É_ACHADO_REAL")) in SIM | NAO | {"duvidoso"}]
    if not marcadas:
        print(f"\n⚠️  A planilha ({fonte}) tem {len(linhas)} candidatos e NENHUM marcado.")
        print("   Sem a sua marcação não há veredito — 'candidato' não é 'achado'.")
        print("   Marque a coluna É_ACHADO_REAL (s/n/duvidoso) e rode de novo.")
        return 1

    st = defaultdict(lambda: {"docs": set(), "reais": 0, "falsos": 0, "duvid": 0,
                              "peso": 0, "excl_reais": 0, "cat": defaultdict(int)})
    for l in marcadas:
        m = str(l.get("modelo") or "?")
        s = st[m]
        s["docs"].add(l.get("documento"))
        v = _v(l.get("É_ACHADO_REAL"))
        if v in SIM:
            s["reais"] += 1
            s["peso"] += PESO.get(_v(l.get("IMPORTA")), 1)
            s["cat"][l.get("categoria")] += 1
            if _v(l.get("EXCLUSIVO")) == "sim":
                s["excl_reais"] += 1
        elif v in NAO:
            s["falsos"] += 1
        else:
            s["duvid"] += 1

    print("\n" + "═" * 84)
    print(f" QUEM ENXERGA — {len(marcadas)} candidatos julgados por Dr. Eduardo")
    print("═" * 84)
    print(f"{'MODELO':22} {'docs':>5} {'REAIS':>6} {'/doc':>6} {'FALSOS':>7} {'invenção':>9} "
          f"{'EXCL.REAIS':>11} {'peso':>6}")
    print("─" * 84)
    rank = []
    for m, s in st.items():
        n = max(len(s["docs"]), 1)
        julg = s["reais"] + s["falsos"]
        inv = (100 * s["falsos"] / julg) if julg else 0
        rank.append((s["excl_reais"] / n, s["reais"] / n, inv, m, s, n))
    for _, _, inv, m, s, n in sorted(rank, reverse=True):
        alerta = " ⚠️" if inv >= 15 else ""
        print(f"{m:22} {n:>5} {s['reais']:>6} {s['reais']/n:>6.1f} {s['falsos']:>7} "
              f"{inv:>8.0f}%{alerta} {s['excl_reais']:>11} {s['peso']:>6}")
    print("═" * 84)

    # rótulo curto e DISTINGUÍVEL: 'claude-sonnet-5' e 'claude-opus-4-8' começam igual —
    # cortar no 1º hífen dava duas colunas "claude" e a tabela virava adivinhação (bug 02/Ago).
    def _curto(m):
        p = m.split("-")
        return (p[1][:4] + p[2][:2]) if len(p) >= 3 else m[:6]

    print("\nPOR CATEGORIA DE ACHADO (só os REAIS):")
    modelos = sorted(st)
    cats = sorted({c for s in st.values() for c in s["cat"]})
    print(f"  {'categoria':26} " + " ".join(f"{_curto(m):>8}" for m in modelos))
    for c in cats:
        print(f"  {str(c)[:26]:26} " + " ".join(f"{st[m]['cat'].get(c,0):>8}" for m in modelos))
    print(f"  {'':26} " + " ".join(f"{'─'*8}" for _ in modelos))
    print(f"  {'legenda':26} " + " ".join(f"{_curto(m):>8}" for m in modelos)
          + "\n      " + " · ".join(f"{_curto(m)}={m}" for m in modelos))

    inventores = [m for m, s in st.items()
                  if s["falsos"] and 100 * s["falsos"] / max(s["reais"] + s["falsos"], 1) >= 15]
    print("\n" + "─" * 84)
    if inventores:
        print(" ⚠️  ATENÇÃO — taxa de invenção ≥15% em: " + ", ".join(inventores))
        print("     Achado falso é pior que achado nenhum: destrói a confiança no produto inteiro.")
        print("     Este modelo NÃO deve ser padrão, por mais achados reais que produza.")
    else:
        print(" Nenhum modelo passou de 15% de invenção nos candidatos julgados.")
    faltam = len(linhas) - len(marcadas)
    if faltam:
        print(f" ({faltam} candidatos ainda sem marcação — o veredito melhora conforme você marca.)")
    print("─" * 84 + "\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
