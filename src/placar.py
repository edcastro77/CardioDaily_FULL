"""
placar.py — PROVA·3 · o PLACAR do experimento (31/Jul/2026).

Lê  outputs/PROVA/prova_bruta.csv  (o que cada modelo respondeu, em cada rodada)
 +  outputs/PROVA/gabarito.xlsx|csv (a coluna CORRETO preenchida pelo Dr. Eduardo)

e responde as QUATRO perguntas que o CardioDaily nunca conseguiu responder:

  0. LINHA DE BASE — quanto o classificador de HOJE acerta? (o número que faltava nos 11 erros)
  1. ACURÁCIA      — quanto cada modelo acerta contra o gabarito?
  2. REPETIBILIDADE— o mesmo modelo responde igual nas N rodadas?
  3. CONCORDÂNCIA  — os modelos concordam entre si? Onde discordam?

E entrega a lista de DISCORDÂNCIA: os artigos em que os modelos não fecharam. Essa lista é a
proposta de porta para REVISAO_HUMANA — o sistema passando a saber quando NÃO sabe, que é
exatamente o que faltou quando 11 erros passaram sem ninguém ver.

O placar roda SEM o gabarito também: aí mede só repetibilidade e concordância, e avisa que
acurácia não foi medida. Confiabilidade não é validade — o placar não deixa confundir as duas.

Uso: python src/placar.py
"""
import os
import csv
import sys
from collections import defaultdict, Counter

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)
SAIDA = os.path.join(_ROOT, "outputs", "PROVA")
CSV_BRUTO = os.path.join(SAIDA, "prova_bruta.csv")


# O Dr. Eduardo preencheu o gabarito em 31/07 do jeito NATURAL: a coluna "CORRETO" recebeu
# "y" (estava certo) ou "errado", e o tipo verdadeiro foi escrito em OBS, em português corrido.
# O programa se adapta a ele, não o contrário. As duas formas de preencher valem:
#   (a) CORRETO = o tipo canônico  → usa direto
#   (b) CORRETO = y/sim/ok         → o certo é o que o classificador disse
#   (c) CORRETO = errado/n/não     → o certo é o que estiver em OBS (normalizado abaixo)
_SIM = {"y", "s", "sim", "ok", "certo", "correto", "true", "1", "x", "v"}
_NAO = {"n", "nao", "não", "errado", "erro", "e", "false", "0"}
_TIPOS_CANON = {"artigo_original", "revisao_sistematica_meta_analise", "revisao_geral",
                "guideline", "ponto_de_vista", "minirevisao", "relato_de_caso",
                "carta_de_pesquisa", "descarte", "duplicata"}


def _normalizar_tipo(texto):
    """'guideline - statement' → guideline · 'revisao sistematica - custoefetividade' →
    revisao_sistematica_meta_analise · 'editorial' → ponto_de_vista. Devolve '' se não entender."""
    t = (texto or "").strip().lower()
    if not t:
        return ""
    t = t.replace("ç", "c").replace("ã", "a").replace("á", "a").replace("é", "e").replace("í", "i")
    if t in _TIPOS_CANON:
        return t
    if "meta" in t or "sistemat" in t:
        return "revisao_sistematica_meta_analise"
    if "guideline" in t or "diretriz" in t or "statement" in t or "consenso" in t:
        return "guideline"
    if "editorial" in t or "viewpoint" in t or "ponto de vista" in t or "coment" in t:
        return "ponto_de_vista"
    if "minirev" in t or "mini rev" in t or "opiniao" in t:
        return "minirevisao"
    if "original" in t:
        return "artigo_original"
    if "revis" in t:
        return "revisao_geral"
    if "caso" in t:
        return "relato_de_caso"
    if "carta" in t or "letter" in t:
        return "carta_de_pesquisa"
    if "descart" in t:
        return "descarte"
    return ""


def _resolver(correto, obs, classificador_disse):
    """Aplica as três formas de preenchimento. Devolve (tipo, motivo_se_nao_deu)."""
    c = (str(correto or "")).strip().lower()
    if not c:
        return "", "vazio"
    direto = _normalizar_tipo(c)
    if c in _SIM:
        return classificador_disse.strip().lower(), ""
    if c in _NAO or c.startswith("err"):
        t = _normalizar_tipo(obs)
        return (t, "") if t else ("", f"marcado errado mas OBS não diz o tipo ({obs!r})")
    if direto:
        return direto, ""
    return "", f"não entendi o valor {correto!r}"


def carregar_gabarito():
    """Devolve ({arquivo: tipo_correto}, fonte, [pendências])."""
    gab, pend = {}, []
    xlsx = os.path.join(SAIDA, "gabarito.xlsx")
    csv_p = os.path.join(SAIDA, "gabarito.csv")

    def _consumir(arquivo, correto, obs, disse):
        arquivo = (arquivo or "").strip()
        if not arquivo:
            return
        t, motivo = _resolver(correto, obs, disse or "")
        if t:
            gab[arquivo] = _eq(t)     # aplica a mesma equivalência ao gabarito
        elif motivo != "vazio":
            pend.append((arquivo, motivo))

    if os.path.exists(xlsx):
        try:
            from openpyxl import load_workbook
            ws = load_workbook(xlsx, data_only=True).active
            cab = [c.value for c in ws[1]]
            ix = {n: (cab.index(n) if n in cab else None)
                  for n in ("arquivo", "CORRETO", "OBS", "classificador_disse")}
            for row in ws.iter_rows(min_row=2, values_only=True):
                _consumir(row[ix["arquivo"]],
                          row[ix["CORRETO"]] if ix["CORRETO"] is not None else "",
                          row[ix["OBS"]] if ix["OBS"] is not None else "",
                          row[ix["classificador_disse"]] if ix["classificador_disse"] is not None else "")
            return gab, "gabarito.xlsx", pend
        except Exception as e:
            print(f"  (não li o xlsx: {type(e).__name__}; tentando o csv)")
    if os.path.exists(csv_p):
        with open(csv_p, encoding="utf-8-sig") as fh:
            for l in csv.DictReader(fh):
                _consumir(l.get("arquivo"), l.get("CORRETO"), l.get("OBS"),
                          l.get("classificador_disse"))
        return gab, "gabarito.csv", pend
    return {}, None, pend


# ─────────────── EQUIVALÊNCIA (decisão do Dr. Eduardo, 31/07) ───────────────
# "ponto de vista e revisão eu deixei como iguais". E não é só opinião dele: o próprio
# classificador_ouro.py, linha 234, faz `if destino == "ponto_de_vista": destino = "minirevisao"`.
# Ou seja, editorial e minirevisão CAEM NA MESMA PASTA e seguem a MESMA trilha.
# Contar isso como erro é medir contra uma taxonomia que o sistema não usa — foi o que eu fiz na
# 1ª conta e por isso o placar disse que os modelos eram piores do que são.
EQUIVALENTES = {
    "ponto_de_vista": "minirevisao",
}


def _eq(t):
    return EQUIVALENTES.get((t or "").strip().lower(), (t or "").strip().lower())


def _barra(p, larg=24):
    cheio = int(round(p * larg))
    return "█" * cheio + "·" * (larg - cheio)


def main():
    if not os.path.exists(CSV_BRUTO):
        print(f"Não achei {CSV_BRUTO} — rode antes: python src/prova_classificador.py"); return 1

    linhas = [l for l in csv.DictReader(open(CSV_BRUTO, encoding="utf-8-sig"))]
    if not linhas:
        print("prova_bruta.csv está vazio."); return 1
    gab, fonte_gab, pendencias = carregar_gabarito()
    if pendencias:
        print(f"\n⚠️  {len(pendencias)} linha(s) do gabarito que eu NÃO consegui interpretar "
              f"(ficam DE FORA da conta — não vou chutar):")
        for arq, motivo in pendencias[:10]:
            print(f"     {arq[:56]:58} {motivo}")

    # resp[arquivo][modelo] = [tipo da rodada 1, rodada 2, ...]
    # Se o CSV tem mais de uma versão de prompt, o placar julga a MAIS NOVA e mostra a anterior
    # ao lado — é assim que se sabe se mexer no prompt melhorou ou piorou.
    versoes = sorted({(l.get("prompt") or "v1") for l in linhas})
    alvo = versoes[-1]
    if len(versoes) > 1:
        print(f"\n  (o CSV tem os prompts {' e '.join(versoes)} — o placar abaixo é do {alvo})")
    linhas_alvo = [l for l in linhas if (l.get("prompt") or "v1") == alvo]

    resp = defaultdict(lambda: defaultdict(list))
    disse_hoje, sem_prova, erros = {}, Counter(), Counter()
    for l in linhas_alvo:
        a, m, t = l["arquivo"], l["modelo"], _eq(l["tipo"])
        resp[a][m].append(t)
        disse_hoje[a] = _eq(l["classificador_disse"])
        if l["tipo"] in ("ERRO", "PARSE_FALHOU"):
            erros[m] += 1
        elif l.get("tem_prova") == "NAO":
            sem_prova[m] += 1
    modelos = sorted({l["modelo"] for l in linhas_alvo})
    linhas = linhas_alvo
    artigos = sorted(resp)
    rodadas = max(len(v) for a in resp for v in resp[a].values())

    print("\n" + "═" * 74)
    print(f" PLACAR DA PROVA · {len(artigos)} artigo(s) · {len(modelos)} modelo(s) · {rodadas} rodada(s)")
    print("═" * 74)

    # ── 0 · LINHA DE BASE: o classificador de hoje ──
    if gab:
        comuns = [a for a in artigos if a in gab]
        acertos_hoje = sum(disse_hoje[a] == gab[a] for a in comuns)
        p = acertos_hoje / len(comuns) if comuns else 0
        print(f"\n0 · LINHA DE BASE — o classificador de HOJE (gabarito: {fonte_gab})")
        print(f"    {_barra(p)}  {acertos_hoje}/{len(comuns)}  ({100*p:.0f} %)")
    else:
        comuns = []
        print("\n0 · LINHA DE BASE — não medida: a coluna CORRETO do gabarito está VAZIA.")
        print("    ⚠️  Sem gabarito NÃO existe acurácia. O que vem abaixo mede só ESTABILIDADE")
        print("        e CONCORDÂNCIA — e um modelo pode ser estável e concordante e estar ERRADO.")

    # ── 1 · ACURÁCIA por modelo (usa a moda das rodadas) ──
    def moda(v):
        return Counter(v).most_common(1)[0][0] if v else ""

    if comuns:
        print(f"\n1 · ACURÁCIA contra o gabarito ({len(comuns)} artigo(s) conferido(s))")
        rank = []
        for m in modelos:
            ok = sum(moda(resp[a].get(m, [])) == gab[a] for a in comuns)
            rank.append((ok / len(comuns), m, ok))
        for p, m, ok in sorted(rank, reverse=True):
            print(f"    {_barra(p)}  {ok:>3}/{len(comuns)}  {100*p:5.1f} %   {m}")

    # ── 2 · REPETIBILIDADE ──
    if rodadas > 1:
        print(f"\n2 · REPETIBILIDADE — o modelo responde IGUAL nas {rodadas} rodadas?")
        for m in modelos:
            iguais = tot = 0
            for a in artigos:
                v = resp[a].get(m, [])
                if len(v) > 1:
                    tot += 1
                    iguais += len(set(v)) == 1
            p = iguais / tot if tot else 0
            print(f"    {_barra(p)}  {iguais:>3}/{tot}  {100*p:5.1f} %   {m}")
    else:
        print("\n2 · REPETIBILIDADE — não medida (rodada única). Use --rodadas 3.")

    # ── 3 · CONCORDÂNCIA entre modelos ──
    print("\n3 · CONCORDÂNCIA entre os modelos (moda de cada um)")
    fechado = divergente = 0
    lista_div = []
    for a in artigos:
        votos = [(m, moda(resp[a][m])) for m in modelos if resp[a].get(m)]
        if len(votos) < 2:
            continue
        if len({v for _, v in votos}) == 1:
            fechado += 1
        else:
            divergente += 1
            lista_div.append((a, votos))
    tot = fechado + divergente
    if tot:
        p = fechado / tot
        print(f"    {_barra(p)}  {fechado}/{tot} unânimes  ({100*p:.0f} %) · {divergente} com divergência")

    # ── qualidade da resposta ──
    if sum(sem_prova.values()) or sum(erros.values()):
        print("\n   Respostas problemáticas:")
        for m in modelos:
            if sem_prova[m] or erros[m]:
                print(f"     {m:28} sem frase de prova: {sem_prova[m]:>3} · erro/parse: {erros[m]:>3}")

    # ── a lista que vira REVISAO_HUMANA ──
    if lista_div:
        print(f"\n4 · ONDE OS MODELOS NÃO FECHARAM ({len(lista_div)}) — candidatos a REVISAO_HUMANA")
        for a, votos in lista_div[:25]:
            marca = ("   ⇒ gabarito=" + gab[a]) if a in gab else ""
            quem = " · ".join(f"{m.split('-')[0]}={v}" for m, v in votos)
            print(f"     {a[:50]:52} {quem}{marca}")
        if len(lista_div) > 25:
            print(f"     … e mais {len(lista_div) - 25}")

    # ── os erros que sobram, contra o gabarito ──
    if comuns:
        falhas = defaultdict(list)
        for a in comuns:
            for m in modelos:
                mm = moda(resp[a].get(m, []))
                if mm and mm != gab[a]:
                    falhas[a].append(f"{m.split('-')[0]}:{mm}")
        if falhas:
            print(f"\n5 · ARTIGOS EM QUE ALGUM MODELO ERROU ({len(falhas)})")
            for a, ms in list(falhas.items())[:25]:
                print(f"     {a[:52]:54} certo={gab[a]:32} {' · '.join(ms)}")

    print("\n" + "═" * 74)
    if not gab:
        print(" VEREDITO: NÃO HÁ VEREDITO. Sem a coluna CORRETO preenchida, nada aqui prova acerto.")
    else:
        print(" Critério do Dr. Eduardo: o modelo MAIS BARATO que bater 100 % de acurácia")
        print(" E 100 % de repetibilidade vence. Empate técnico → fica o mais barato.")
    print("═" * 74 + "\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
