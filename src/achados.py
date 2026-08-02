"""
achados.py — O INSTRUMENTO QUE FALTAVA (02/Ago/2026).

POR QUE EXISTE
--------------
Escolhi o gpt-5.6-terra como padrão da perícia medindo TABELAS, TEMPO e CUSTO — porque era o que
eu conseguia contar. Aí o claude-sonnet-5, na meta-análise do sirolimus, achou isto:

    "O texto metodológico declara: 'Publication bias was not formally assessed...'. Entretanto,
     a Tabela 2 apresenta teste de Egger para todos os nove desfechos (TLF t=−0,67, p=0,54).
     Ou seja, o teste foi de fato calculado, contradizendo a afirmação de que não foi feito."

Nenhum outro modelo viu. E isso — não a contagem de tabelas — é o que o assinante paga para receber.
**Eu estava otimizando o que era fácil medir.** Este programa mede o que importa.

O QUE ELE FAZ
-------------
1. Lê os comparativos (o .md que o painel já gera) e separa a perícia de cada modelo.
2. Extrai os CANDIDATOS a achado de perícia — frases onde o modelo confronta o artigo, não o resume:
   contradição interna, número que não fecha, ausência de método, subgrupo pescado, extrapolação.
3. Marca quais achados são EXCLUSIVOS de um modelo (o dado que decide a escolha).
4. Gera a planilha para o Dr. Eduardo julgar: é achado de verdade? importa?

O QUE ELE **NÃO** FAZ: não chama LLM, não usa rede, não gasta um centavo. É leitura de texto.
Julgar se o achado é verdadeiro é do perito — a máquina não tem como saber.

Uso:  python src/achados.py                    (todos os comparativos em outputs/PROVA_PROMPTS/)
      python src/achados.py <arquivo.md> ...
"""
import os
import re
import sys
import csv
import glob

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)
SAIDA = os.path.join(_ROOT, "outputs", "PROVA_PROMPTS")

MODELOS = re.compile(r"claude-sonnet-5|claude-opus-4-8|gpt-5\.6-sol|gpt-5\.6-terra|gpt-5\.6-luna")

# ─────────────── O QUE É "ACHADO DE PERÍCIA" ───────────────
# Não é qualquer frase crítica. É a frase em que o modelo CONFRONTA o artigo com ele mesmo, ou com
# o que o método exigiria. Cada padrão abaixo veio de uma perícia real lida em 01–02/Ago.
PADROES = [
    ("CONTRADIÇÃO INTERNA",
     r"contradiz|contradi[çc][ãa]o|inconsist[êe]nc|discrep[âa]nc|n[ãa]o bate|conflita com|"
     r"entretanto,? a [Tt]abela|por[ée]m a [Tt]abela|mas a [Tt]abela|difere do que|"
     r"declara.{0,80}(entretanto|por[ée]m|no entanto)"),
    ("MÉTODO AUSENTE",
     r"n[ãa]o (foi|houve|apresenta|reporta|realizou|testou|avaliou|declara)\w*\s"
     r"(teste|an[áa]lise|avalia|busca|prot[óo]colo|registro|meta-regress|sensibilidade|"
     r"cegamento|duplicata|heterogeneidade|vi[ée]s)|sem meta-regress|sem an[áa]lise de sensibilidade|"
     r"sem prot[óo]colo|sem registro pr[ée]vio|n[ãa]o pr[ée]-especificad"),
    ("NÚMERO FRÁGIL",
     r"poucos eventos|baixo poder|IC.{0,30}(amplo|cruza|inclui o nulo)|"
     r"depende de (um|1|dois|2) [úu]nico|dominad[oa] por (um|1) [úu]nico|"
     r"sumiria|deixaria de ser significativ|fr[áa]gil"),
    ("EXTRAPOLAÇÃO",
     r"vai(?:am)? al[ée]m d[oa]|extrapola|n[ãa]o decorre|n[ãa]o sustenta|"
     r"associa[çc][ãa]o.{0,40}causa|generaliza[çc][ãa]o indevida|"
     r"apresentad[oa] como.{0,40}(sem|mas n[ãa]o)"),
    ("SUBGRUPO / DESFECHO",
     r"fishing|pesca|post[- ]hoc|desfecho (substituto|surrogate|composto).{0,60}"
     r"(mistura|carreg|dominad)|trocad[oa] ap[óo]s|mudan[çc]a de desfecho|"
     r"intera[çc][ãa]o.{0,30}n[ãa]o (era |foi )?pr[ée]-especificad"),
    ("CONFLITO DE INTERESSE",
     r"patrocinad|financiad[oa] (integralmente|pela ind[úu]stria)|conflito de interesse.{0,60}"
     r"(n[ãa]o|todos|maioria)|v[íi]nculo com (a|o) fabricante"),
]
PADROES = [(nome, re.compile(p, re.I)) for nome, p in PADROES]

# frases-clichê que PARECEM achado e não são: o modelo só repetindo a instrução do prompt
RUIDO = re.compile(r"^(##|\||-{3,}|\*{3,})|"
                   r"(a lei do n[úu]mero|conforme a instru|de acordo com o prompt|"
                   r"como pedido|nesta se[çc][ãa]o|a seguir|nota de aplicabilidade|nota de rigor)", re.I)


def _blocos(caminho):
    """Separa o texto por modelo. Devolve {modelo: texto}."""
    t = open(caminho, encoding="utf-8", errors="ignore").read()
    marcas = [(m.start(), m.group(0)) for m in MODELOS.finditer(t)]
    blocos = {}
    for i, (p, nome) in enumerate(marcas):
        fim = marcas[i + 1][0] if i + 1 < len(marcas) else len(t)
        if nome not in blocos:                       # a 1ª ocorrência é o cabeçalho da aba
            blocos[nome] = t[p:fim]
    return blocos


def _frases(texto):
    """Quebra em frases, ignorando tabela e cabeçalho."""
    linhas = [l for l in texto.splitlines() if not RUIDO.match(l.strip())]
    corrido = " ".join(linhas)
    return [f.strip() for f in re.split(r"(?<=[.!?])\s+(?=[A-ZÀ-Ú*])", corrido) if 60 <= len(f.strip()) <= 700]


def extrair(texto):
    """Devolve [(categoria, frase)] — os candidatos a achado de perícia."""
    out = []
    for f in _frases(texto):
        for nome, rx in PADROES:
            if rx.search(f):
                out.append((nome, re.sub(r"\s+", " ", f)))
                break                                 # uma categoria por frase, a primeira que casa
    return out


def _assinatura(frase):
    """Chave grosseira p/ dizer se dois modelos acharam A MESMA coisa: números + palavras longas."""
    nums = set(re.findall(r"\d+[,.]?\d*", frase))
    pal = {w.lower() for w in re.findall(r"[A-Za-zÀ-ú]{7,}", frase)}
    return nums, pal


def _parecido(a, b, limiar=0.45):
    na, pa = _assinatura(a)
    nb, pb = _assinatura(b)
    if na and nb and len(na & nb) >= 2:               # dois números iguais = mesmo achado
        return True
    if not pa or not pb:
        return False
    return len(pa & pb) / min(len(pa), len(pb)) >= limiar


def comparar(caminho):
    """Devolve as linhas da planilha para UM comparativo."""
    blocos = _blocos(caminho)
    por_modelo = {m: extrair(b) for m, b in blocos.items()}
    doc = os.path.basename(caminho).replace("_comparativo", "").replace(".md", "")[:60]

    linhas = []
    for modelo, achados in por_modelo.items():
        for cat, frase in achados:
            # quem MAIS achou isto?
            tambem = [outro for outro, lista in por_modelo.items()
                      if outro != modelo and any(_parecido(frase, f2) for _, f2 in lista)]
            linhas.append({
                "documento": doc,
                "modelo": modelo,
                "categoria": cat,
                "EXCLUSIVO": "SIM" if not tambem else "",
                "tambem_acharam": " · ".join(sorted(tambem)),
                "achado": frase[:600],
                "É_ACHADO_REAL": "",      # ← Dr. Eduardo: s / n
                "IMPORTA": "",            # ← Dr. Eduardo: alto / medio / baixo
                "OBS": "",
            })
    return linhas, por_modelo


def main(arquivos):
    if not arquivos:
        arquivos = glob.glob(os.path.join(_ROOT, "outputs", "**", "*comparativo*.md"), recursive=True)
    # DEDUP por caminho REAL: o glob recursivo já inclui PROVA_PROMPTS, e o mesmo arquivo entrando
    # duas vezes dobra a contagem de documentos (bug pego em 02/Ago — 7 arquivos viravam 14 "docs").
    arquivos = sorted({os.path.realpath(a) for a in arquivos})
    if not arquivos:
        print("Nenhum comparativo .md encontrado.\n"
              f"Ponha os arquivos em {SAIDA}/ ou passe os caminhos como argumento."); return 1

    todas, resumo = [], {}
    for f in arquivos:
        linhas, por_modelo = comparar(f)
        # Comparativo de RECUSA (veredito vazio) não é documento analisado — não entra no denominador.
        uteis = {m: a for m, a in por_modelo.items() if len(_blocos(f).get(m, "")) > 2000}
        if not uteis:
            print(f"  {os.path.basename(f)[:52]:54}   — (só recusas, fora da conta)")
            continue
        todas += [l for l in linhas if l["modelo"] in uteis]
        for m in uteis:
            r = resumo.setdefault(m, {"achados": 0, "exclusivos": 0, "docs": 0})
            r["achados"] += len(por_modelo[m])
            r["docs"] += 1
        for l in linhas:
            if l["EXCLUSIVO"] == "SIM" and l["modelo"] in uteis:
                resumo[l["modelo"]]["exclusivos"] += 1
        print(f"  {os.path.basename(f)[:52]:54} {len([l for l in linhas if l['modelo'] in uteis]):>3} candidatos")

    if not todas:
        print("\nNenhum candidato a achado encontrado — confira se os arquivos têm as perícias."); return 1

    os.makedirs(SAIDA, exist_ok=True)
    campos = list(todas[0].keys())
    csv_p = os.path.join(SAIDA, "achados_para_marcar.csv")
    with open(csv_p, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=campos)
        w.writeheader(); w.writerows(todas)

    xlsx_p = None
    try:
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook(); ws = wb.active; ws.title = "ACHADOS"
        ws.append(campos)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="0B3D91")
        for l in todas:
            ws.append([l[c] for c in campos])
        for nome, opcoes in (("É_ACHADO_REAL", '"s,n,duvidoso"'), ("IMPORTA", '"alto,medio,baixo"')):
            col = ws.cell(row=1, column=campos.index(nome) + 1).column_letter
            dv = DataValidation(type="list", formula1=opcoes, allow_blank=True)
            ws.add_data_validation(dv); dv.add(f"{col}2:{col}{len(todas)+1}")
        larg = {"documento": 34, "modelo": 20, "categoria": 22, "EXCLUSIVO": 11,
                "tambem_acharam": 26, "achado": 110, "É_ACHADO_REAL": 15, "IMPORTA": 11, "OBS": 24}
        for c in ws[1]:
            ws.column_dimensions[c.column_letter].width = larg.get(c.value, 12)
        for row in ws.iter_rows(min_row=2):
            row[campos.index("achado")].alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "D2"
        xlsx_p = os.path.join(SAIDA, "achados_para_marcar.xlsx")
        wb.save(xlsx_p)
    except ImportError:
        print("  (sem openpyxl — só o CSV)")

    print("\n" + "═" * 72)
    print(f"{'MODELO':22} {'docs':>5} {'candidatos':>11} {'EXCLUSIVOS':>11}  (por doc)")
    print("─" * 72)
    for m, r in sorted(resumo.items(), key=lambda x: -x[1]["exclusivos"]):
        print(f"{m:22} {r['docs']:>5} {r['achados']:>11} {r['exclusivos']:>11}  "
              f"{r['exclusivos']/max(r['docs'],1):>8.1f}")
    print("═" * 72)
    print(f"\n→ {xlsx_p or csv_p}")
    print("   Marque É_ACHADO_REAL (s/n/duvidoso) e IMPORTA (alto/medio/baixo).")
    print("   Depois: python src/placar_achados.py")
    print("\n   ⚠️  'EXCLUSIVOS' é candidato, NÃO é acerto. Um modelo pode inventar contradição que")
    print("       não existe — e isso seria PIOR que não achar nada. Quem decide é você, lendo.")
    return 0


if __name__ == "__main__":
    sys.exit(main([os.path.expanduser(a) for a in sys.argv[1:]]))
