"""
gabarito.py — PROVA·1 · gera o PADRÃO-OURO do classificador (31/Jul/2026).

POR QUE ISTO EXISTE (a lição dos 11 erros em 62 artigos):
  O classificador nunca teve padrão de referência. Sem referência, ninguém sabe se ele acerta 90 %
  ou 60 %, e NENHUM conserto pode ser provado como conserto — vira opinião contra opinião.
  Aqui o Dr. Eduardo diz, de uma vez, o que cada artigo É. Daí em diante toda mudança no
  classificador é medida contra isto: bate o gabarito ou não entra.

O QUE ELE FAZ: lista o que está HOJE em CLASSIFICADOS/, com a pasta em que o classificador o pôs,
o DOI, e — o que economiza o tempo do Dr. Eduardo — O RÓTULO QUE O PRÓPRIO ARTIGO IMPRIME
(ORIGINAL RESEARCH ARTICLE, AHA SCIENTIFIC STATEMENT, THE HEART OF THE MATTER…), lido das
páginas 1–3. A coluna CORRETO vem VAZIA: quem preenche é ele.

O QUE ELE **NÃO** FAZ (garantia): não move arquivo, não renomeia, não chama LLM, não usa rede,
não toca no Supabase. É somente leitura. Os errados continuam nas pastas erradas — de propósito:
é a planilha que vira a lista de correção.

Uso:  python src/gabarito.py [--tudo]      (--tudo inclui _PUBLICADOS e _RECUSADOS)
Saída: outputs/PROVA/gabarito.csv  (+ .xlsx com lista suspensa, se houver openpyxl)
"""
import os
import re
import sys
import csv
import glob

import fitz

_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_HERE)
BASE = os.path.join(_ROOT, "ARTIGOS", "CLASSIFICADOS")
SAIDA = os.path.join(_ROOT, "outputs", "PROVA")

# pasta no disco -> rótulo canônico (o vocabulário que o Dr. Eduardo vai usar na coluna CORRETO)
PASTA_TIPO = {
    "ARTIGOS_ORIGINAIS": "artigo_original",
    "META_ANALISES": "revisao_sistematica_meta_analise",
    "REVISOES": "revisao_geral",
    "GUIDELINES": "guideline",
    "EDITORIAIS": "ponto_de_vista",
    "MINIRREVISOES": "minirevisao",
}
# DECISÃO D-01 do Dr. Eduardo (31/07): revisão sistemática = meta-análise (mesma trilha).
TIPOS_VALIDOS = list(dict.fromkeys(PASTA_TIPO.values())) + ["DESCARTE", "DUPLICATA"]

# O rótulo que a REVISTA imprime. Medido em 158 PDFs (Circulation + ESC): sai em TEXTO, não é figura.
_ROTULO = re.compile(
    r"(ORIGINAL RESEARCH ARTICLE|ORIGINAL RESEARCH|ORIGINAL ARTICLE|ORIGINAL INVESTIGATION"
    r"|CLINICAL RESEARCH|RESEARCH ARTICLE|BRIEF REPORT|RESEARCH LETTER"
    r"|STATE[- ]OF[- ]THE[- ]ART|AHA SCIENTIFIC STATEMENT|SCIENTIFIC STATEMENT"
    r"|CLINICAL PRACTICE GUIDELINE|PRACTICE GUIDELINE|GUIDELINE"
    r"|SYSTEMATIC REVIEW AND META[- ]ANALYSIS|META[- ]ANALYSIS|SYSTEMATIC REVIEW"
    r"|REVIEW ARTICLE|THE HEART OF THE MATTER|IN DEPTH|FRONTIERS|ISSUE @ A GLANCE"
    r"|EDITORIAL COMMENT|EDITORIAL|VIEWPOINT|PERSPECTIVE|COMMENTARY"
    r"|SPECIAL REPORT|CONSENSUS DOCUMENT|CONSENSUS STATEMENT|POSITION PAPER|CASE REPORT)", re.I)

_DOI = re.compile(r"10\.\d{4,9}/[-._;()/:A-Za-z0-9]+")

# PDF de revista traz caractere de controle no texto (o Circulation põe \x08 no fim da linha do DOI).
# O Excel RECUSA esses bytes (IllegalCharacterError) — limpar aqui, não na hora de gravar.
_CTRL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def _limpo(s):
    return _CTRL.sub("", s or "").strip()


def _paginas(caminho, n=3):
    """Texto das n primeiras páginas + o da 1ª isolado. (Medido: página 1 sozinha é capa em 77 % do ESC.)"""
    doc = fitz.open(caminho)
    p1 = doc[0].get_text() if len(doc) else ""
    p13 = "".join(doc[i].get_text() for i in range(min(n, len(doc))))
    return p1, p13, len(doc)


def _linhas_uteis(texto, quantas=3):
    """Primeiras linhas com conteúdo — pula boilerplate de licença/download que abre PDF de MDPI/Oxford."""
    lixo = re.compile(r"^(downloaded from|copyright|©|received:|revised:|accepted:|published:|"
                      r"academic editor|licensee|this article is an open access|https?://)", re.I)
    out = []
    for l in (x.strip() for x in texto.splitlines()):
        l = _limpo(l)
        if l and not lixo.match(l) and len(l) > 3:
            out.append(l)
        if len(out) >= quantas:
            break
    return " | ".join(out)


def montar(incluir_arquivados=False):
    pastas = list(PASTA_TIPO)
    if incluir_arquivados:
        pastas += ["_PUBLICADOS", "_RECUSADOS"]

    linhas = []
    for pasta in pastas:
        for f in sorted(glob.glob(os.path.join(BASE, pasta, "*.pdf"))):
            if os.path.basename(f).startswith("._"):
                continue
            try:
                p1, p13, npg = _paginas(f)
            except Exception as e:
                p1 = p13 = ""
                npg = 0
                print(f"  ⚠️ não abriu: {os.path.basename(f)[:50]} ({type(e).__name__})")
            m = _DOI.search(p13[:8000])
            doi = m.group(0).rstrip(".,;:") if m else ""
            rot_p1 = sorted({x.group(0).upper() for x in _ROTULO.finditer(p1[:1800])})
            rot_p13 = sorted({x.group(0).upper() for x in _ROTULO.finditer(p13[:9000])})
            linhas.append({
                "n": len(linhas) + 1,
                "arquivo": _limpo(os.path.basename(f)),
                "pasta_hoje": pasta,
                "classificador_disse": PASTA_TIPO.get(pasta, pasta),
                "CORRETO": "",                       # ← o Dr. Eduardo preenche
                "OBS": "",                           # ← e comenta, se quiser
                "rotulo_impresso_no_artigo": " / ".join(rot_p13) or "(nenhum)",
                "rotulo_visivel_na_pag1": " / ".join(rot_p1) or "(nenhum)",
                "primeiras_linhas": _linhas_uteis(p13)[:180],
                "doi": _limpo(doi),
                "chars_pag1": len(p1),
                "chars_pag1a3": len(p13),
                "paginas": npg,
            })
    return linhas


def gravar(linhas):
    os.makedirs(SAIDA, exist_ok=True)
    campos = list(linhas[0].keys())
    csv_path = os.path.join(SAIDA, "gabarito.csv")
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=campos)
        w.writeheader()
        w.writerows(linhas)

    xlsx_path = None
    try:                                   # xlsx com lista suspensa: preencher 105 linhas sem digitar
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "GABARITO"
        ws.append(campos)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="0B3D91")
            c.alignment = Alignment(vertical="center")
        for ln in linhas:
            ws.append([ln[c] for c in campos])
        col_correto = campos.index("CORRETO") + 1
        letra = ws.cell(row=1, column=col_correto).column_letter
        dv = DataValidation(type="list", formula1='"' + ",".join(TIPOS_VALIDOS) + '"', allow_blank=True)
        dv.error = "Use um dos tipos da lista."
        ws.add_data_validation(dv)
        dv.add(f"{letra}2:{letra}{len(linhas) + 1}")
        for c in ws[1]:
            ws.column_dimensions[c.column_letter].width = \
                {"arquivo": 52, "CORRETO": 32, "classificador_disse": 30, "pasta_hoje": 20,
                 "rotulo_impresso_no_artigo": 34, "rotulo_visivel_na_pag1": 26,
                 "primeiras_linhas": 60, "doi": 28, "OBS": 24}.get(c.value, 12)
        ws.freeze_panes = "C2"
        xlsx_path = os.path.join(SAIDA, "gabarito.xlsx")
        wb.save(xlsx_path)
    except ImportError:
        print("  (openpyxl ausente — só o CSV foi gerado; abre no Excel do mesmo jeito)")
    return csv_path, xlsx_path


def main():
    tudo = "--tudo" in sys.argv
    if not os.path.isdir(BASE):
        print(f"Não achei {BASE}"); return 1
    print(f"\nGABARITO · lendo {'TUDO (inclui arquivados)' if tudo else 'as pastas de tipo'} — somente leitura\n")
    linhas = montar(tudo)
    if not linhas:
        print("Nenhum PDF encontrado."); return 1
    csv_p, xlsx_p = gravar(linhas)

    por_pasta = {}
    sem_rotulo = 0
    for l in linhas:
        por_pasta[l["pasta_hoje"]] = por_pasta.get(l["pasta_hoje"], 0) + 1
        sem_rotulo += l["rotulo_impresso_no_artigo"] == "(nenhum)"
    print(f"{len(linhas)} artigo(s):")
    for k, v in sorted(por_pasta.items()):
        print(f"   {v:>4}  {k}")
    print(f"\nRótulo impresso encontrado em {len(linhas) - sem_rotulo}/{len(linhas)} "
          f"({100 * (len(linhas) - sem_rotulo) // len(linhas)} %) lendo páginas 1–3.")
    print(f"\n→ {xlsx_p or csv_p}")
    print("   Preencha SÓ a coluna CORRETO. Nada foi movido; os errados seguem nas pastas erradas.")
    print(f"   Vocabulário: {' · '.join(TIPOS_VALIDOS)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
