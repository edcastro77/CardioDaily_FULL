"""
lista_dos_retidos.py — a planilha dos RETIDOS, com o MOTIVO ao lado do título.

═══ 22/Ago/2026 — POR QUE ═══
Palavras dele: *"na prática a pasta hoje tem muito mais de 10 artigos que não consigo imaginar
o porquê que foram barrados. Vou separar manualmente todos eles."*

Ele vai fazer à mão, e está certo — a régua é dele. O que NÃO pode é ele abrir 255 PDFs para
descobrir o motivo de cada um. O motivo já existe: está no `_CANONICO.md` de cada pacote, na
linha `delatores`, escrita pelo MOTOR (não pelo LLM). Este programa só junta tudo numa planilha
e põe o motivo na coluna de trás do título.

Sai em `saidas/RETIDOS_para_revisar.xlsx`, com filtro em toda coluna. A coluna **VOLTAR** é
dele: escreva `x` (ou `sim`) na linha do artigo que deve voltar para a fila.

═══ LEI 12 — O QUE VOCÊ MARCAR É SÓ SEU ═══
Este programa **nunca** sobrescreve um arquivo já existente. Se `RETIDOS_para_revisar.xlsx`
já estiver lá, ele grava com sufixo de data e avisa. Marcação a mão é trabalho que só o Dr.
Eduardo consegue refazer — em 20/Ago eu destruí um gabarito dele copiando por cima de um
upload de 0 byte, e a lei nasceu disso.

Depois de marcar, quem devolve à fila é `scripts/devolver_retidos.py` (que também não escreve
na planilha — só lê).

Uso:  python3 scripts/lista_dos_retidos.py
"""
import csv
import datetime
import glob
import json
import os
import re
import sys

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RETIDOS = os.path.join(RAIZ, "ARTIGOS", "CLASSIFICADOS", "_RETIDOS_PELA_REGUA")
SAIDAS = os.path.join(RAIZ, "saidas")

COLUNAS = ["VOLTAR", "nota", "rigor", "tipo", "ano", "revista", "titulo",
           "POR QUE FOI BARRADO (o motor)", "muda_conduta", "arquivo"]


def _pacote(nome_pdf):
    base = os.path.splitext(nome_pdf)[0]
    for padrao in (os.path.join(RAIZ, "outputs", "STAGING", base),
                   os.path.join(RAIZ, "outputs", "ARQUIVO", "*", base)):
        for p in glob.glob(padrao):
            if os.path.isdir(p):
                return p
    return None


def _campo(txt, chave):
    m = re.search(rf'{chave}:\s*"(.*?)"\s*$', txt, re.M)
    return m.group(1).strip() if m else ""


def _num(txt, chave):
    m = re.search(rf"{chave}:\s*(-?\d+)", txt)
    return int(m.group(1)) if m else None


def _delatores(txt):
    """A lista `delatores:` do canônico — em português, uma frase por motivo."""
    m = re.search(r"delatores:\s*\[(.*?)\]", txt, re.S)
    if not m:
        return ""
    itens = re.findall(r'"((?:[^"\\]|\\.)*)"', m.group(1))
    return " · ".join(i.strip() for i in itens if i.strip())


def linhas():
    if not os.path.isdir(RETIDOS):
        print(f"⛔ não achei {RETIDOS}")
        return []
    out = []
    for f in sorted(os.listdir(RETIDOS)):
        if not f.lower().endswith(".pdf"):
            continue
        p = _pacote(f)
        if not p:
            out.append({"VOLTAR": "", "nota": "", "rigor": "", "tipo": "", "ano": "",
                        "revista": "", "titulo": os.path.splitext(f)[0][:120],
                        "POR QUE FOI BARRADO (o motor)":
                            "⚠️ sem pacote no disco — motivo NÃO registrado",
                        "muda_conduta": "", "arquivo": f})
            continue
        can = glob.glob(os.path.join(p, "*_CANONICO.md"))
        txt = open(can[0], encoding="utf-8", errors="ignore").read() if can else ""
        tipo = ""
        for fj in glob.glob(os.path.join(p, "*_fatos.json")):
            try:
                tipo = (json.load(open(fj, encoding="utf-8")) or {}).get("tipo_documento") or ""
            except Exception:
                pass
        out.append({
            "VOLTAR": "",
            "nota": _num(txt, "nota_aplicabilidade_clinica"),
            "rigor": _num(txt, "nota_trabalho_estatistico"),
            "tipo": tipo or _campo(txt, "tipo"),
            "ano": _campo(txt, "ano"),
            "revista": _campo(txt, "revista")[:44],
            "titulo": _campo(txt, "titulo")[:150] or os.path.splitext(f)[0][:120],
            "POR QUE FOI BARRADO (o motor)": _delatores(txt) or "(sem delator registrado)",
            "muda_conduta": _campo(txt, "muda_conduta"),
            "arquivo": f,
        })
    return out


def _alvo(ext):
    base = os.path.join(SAIDAS, f"RETIDOS_para_revisar.{ext}")
    if not os.path.exists(base):
        return base, False
    # LEI 12: NUNCA por cima. Se já existe, é porque ele pode ter marcado.
    carimbo = datetime.datetime.now().strftime("%d%b-%Hh%M")
    return os.path.join(SAIDAS, f"RETIDOS_para_revisar_{carimbo}.{ext}"), True


def main():
    os.makedirs(SAIDAS, exist_ok=True)
    dados = linhas()
    if not dados:
        print("nada em _RETIDOS_PELA_REGUA.")
        return 0

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        alvo, existia = _alvo("xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "retidos"
        ws.append(COLUNAS)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="0B3D91")
            c.alignment = Alignment(vertical="center")
        for d in dados:
            ws.append([d.get(c, "") for c in COLUNAS])
        larg = {"VOLTAR": 9, "nota": 6, "rigor": 7, "tipo": 15, "ano": 6, "revista": 26,
                "titulo": 62, "POR QUE FOI BARRADO (o motor)": 92, "muda_conduta": 16,
                "arquivo": 40}
        for i, c in enumerate(COLUNAS, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = larg.get(c, 18)
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        wb.save(alvo)
    except ImportError:
        alvo, existia = _alvo("csv")
        with open(alvo, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=COLUNAS)
            w.writeheader()
            w.writerows(dados)

    import collections
    por_nota = collections.Counter(d["nota"] for d in dados)
    por_motivo = collections.Counter()
    for d in dados:
        primeiro = str(d["POR QUE FOI BARRADO (o motor)"]).split(" · ")[0][:66]
        por_motivo[primeiro] += 1

    print(f"{len(dados)} retidos\n")
    print("   por NOTA:  " + " · ".join(f"{n}→{q}" for n, q in sorted(
        por_nota.items(), key=lambda x: (x[0] is None, x[0]))))
    print("\n   OS MOTIVOS MAIS COMUNS (1º delator de cada):")
    for m, q in por_motivo.most_common(12):
        print(f"      {q:>4}×  {m}")
    print(f"\n✔ {alvo}")
    if existia:
        print("   ⚠️  já havia um arquivo com o nome padrão — NÃO sobrescrevi (LEI 12).")
    print("\n   Marque `x` na coluna VOLTAR das linhas que devem voltar para a fila.")
    print("   Depois:  python3 scripts/devolver_retidos.py <caminho do arquivo marcado>")
    return 0


if __name__ == "__main__":
    sys.exit(main())
