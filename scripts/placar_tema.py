"""
placar_tema.py — quanto o classificador de tema acerta, medido contra o gabarito DELE.

═══════════════════════════ POR QUE ISTO EXISTE ═══════════════════════════

Em 20/Ago o Dr. Eduardo abriu o Supabase e viu `tema` NULL em 117 de 616 linhas. A causa
imediata era outra (o `marcar_temas.py` é um segundo portão e o publicador nunca soube da
coluna — ver CADERNO, PARTE 19). Mas ao listar as 48 diretrizes apareceu algo pior que o vazio:

    "Guidelines for the Prevention of Work-Related Musculoskeletal Disorders" → Coronária/DAC
    "Perioperative Anemia and Blood Management in Cardiac Surgery"            → Insuficiência Cardíaca
    "Observational Comparative Research in CV and Brain Health"               → Imagem Cardiovascular

**Tema errado é pior que tema vazio.** O vazio você vê; o errado entrega cardiologia do esporte
para quem assinou coronária, e o assinante conclui que o produto não sabe o que está fazendo.

E a precisão do mapa MeSH **nunca foi medida**. Em 17/Ago eu mostrei cobertura (*"499 de 520 com
tema"*) e chamei de resultado — cobertura é quantos foram marcados, não quantos foram marcados
CERTO. É a mesma confusão que a LEI 7 proíbe: relatar o sucesso de um componente como se fosse
o do todo.

═══ O MÉTODO (o mesmo dos 105 artigos do classificador, em 31/Jul) ═══
1. `GABARITO_TEMA_cego.xlsx` — 40 artigos, 20 do MeSH e 20 do LLM, **embaralhados**, com o
   tema do sistema FORA da planilha. Ele marca o que daria, olhando só o título.
2. Este programa compara e devolve o placar, separado por origem.

⚠️ ESTE ARQUIVO FOI ESCRITO ANTES DE ELE MARCAR. É de propósito: placar escrito depois de ver
o resultado é placar ajustado ao resultado.

Uso:  python3 scripts/placar_tema.py
"""
import json
import os
import sys
from collections import Counter, defaultdict

_AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.dirname(_AQUI)
sys.path.insert(0, os.path.join(RAIZ, "src"))

PLANILHA = os.path.join(RAIZ, "saidas", "GABARITO_TEMA_cego.xlsx")
RESPOSTAS = os.path.join(RAIZ, "saidas", "gabarito_tema_RESPOSTAS_DO_SISTEMA.json")


def ler_gabarito():
    """(doc_id → {marcado, secundario, obs}) a partir da planilha que ELE preencheu."""
    from openpyxl import load_workbook
    respostas = json.load(open(RESPOSTAS, encoding="utf-8"))
    # a planilha não tem doc_id (para não virar pista); casa pelo TÍTULO, que é único aqui
    por_titulo = {v["titulo"]: k for k, v in respostas.items()}
    ws = load_workbook(PLANILHA).active
    out, sem_marcar = {}, []
    for r in range(6, ws.max_row + 1):
        titulo = ws.cell(r, 2).value
        if not titulo:
            continue
        doc = por_titulo.get(titulo)
        if doc is None:
            print(f"   ⚠️ linha {r}: título não bate com a amostra — ignorada")
            continue
        marcado = (ws.cell(r, 5).value or "").strip()
        if not marcado:
            sem_marcar.append(titulo[:60]); continue
        out[doc] = {"marcado": marcado,
                    "secundario": (ws.cell(r, 6).value or "").strip(),
                    "obs": (ws.cell(r, 7).value or "").strip()}
    return out, sem_marcar


def main():
    # ⚠️ 20/Ago — ESTE GUARDA NASCEU DE UM ERRO MEU, E ELE APAGOU TRABALHO DELE.
    # Ele marcou as 40 linhas e mandou a planilha. O arquivo chegou com **0 bytes** (o upload
    # não trouxe conteúdo), e eu copiei esse arquivo vazio POR CIMA do que estava em `saidas/`
    # sem conferir o tamanho. As marcações se perderam, e `saidas/` não está no git.
    # É a família de sempre — ausência tratada como dado — só que desta vez destrutiva.
    # Agora: qualquer arquivo suspeito de vazio PARA aqui, antes de qualquer coisa.
    for f in (PLANILHA, RESPOSTAS):
        if not os.path.exists(f):
            print(f"⛔ falta {f}"); return 1
        if os.path.getsize(f) < 1024:
            print(f"⛔ {os.path.basename(f)} tem {os.path.getsize(f)} bytes — está vazio ou "
                  f"truncado. NÃO vou ler nem sobrescrever nada.")
            return 1

    sistema = json.load(open(RESPOSTAS, encoding="utf-8"))
    gab, sem_marcar = ler_gabarito()

    if sem_marcar:
        print(f"⚠️ {len(sem_marcar)} linha(s) sem marcação — ficam de fora da conta:")
        for t in sem_marcar[:5]:
            print(f"      · {t}")
        print()
    if not gab:
        print("⛔ Nenhuma linha marcada. Preencha a coluna E da planilha antes de rodar.")
        return 1

    # ── o placar, separado por origem: é a pergunta que decide o que consertar ──
    acerto = Counter(); total = Counter()
    parcial = Counter()          # errou o principal MAS acertou no secundário
    fora = []                    # ele marcou FORA DO ESCOPO
    erros = defaultdict(list)
    for doc, m in gab.items():
        s = sistema[doc]; origem = s["origem"]
        total[origem] += 1
        if m["marcado"] == "FORA DO ESCOPO":
            fora.append((s["titulo"][:66], s["tema"])); continue
        if m["marcado"] == s["tema"]:
            acerto[origem] += 1
        elif m["marcado"] and m["marcado"] == (s.get("secundario") or ""):
            parcial[origem] += 1
            erros[origem].append((s["titulo"][:58], s["tema"], m["marcado"], "2º"))
        else:
            erros[origem].append((s["titulo"][:58], s["tema"], m["marcado"], ""))

    print("═" * 76)
    print("PLACAR DO TEMA — sistema vs. gabarito do Dr. Eduardo")
    print("═" * 76)
    print(f"{'origem':8} {'n':>4} {'acertou':>9} {'2º tema':>9} {'errou':>7}")
    for o in ("mesh", "llm"):
        if not total[o]:
            continue
        n, a, p = total[o], acerto[o], parcial[o]
        e = n - a - p - len([1 for t, _ in fora if True]) * 0
        print(f"{o:8} {n:>4} {a:>4} ({100*a//n:>3}%) {p:>4}      {n-a-p:>5}")
    N = sum(total.values()); A = sum(acerto.values()); P = sum(parcial.values())
    print("─" * 76)
    print(f"{'TOTAL':8} {N:>4} {A:>4} ({100*A//max(N,1):>3}%) {P:>4}      {N-A-P:>5}")
    print(f"\ncom o 2º tema valendo: {100*(A+P)//max(N,1)}%")

    if fora:
        print(f"\n⛔ {len(fora)} artigo(s) que NÃO deveriam estar no acervo (ele marcou FORA DO ESCOPO):")
        for t, tema in fora:
            print(f"     [{tema}] {t}")

    for o in ("mesh", "llm"):
        if not erros[o]:
            continue
        print(f"\n── onde o {o.upper()} errou ──")
        for t, deu, devia, marca in erros[o]:
            print(f"   {t}")
            print(f"      sistema: {deu:28} você: {devia} {marca}")

    # ── para onde o erro CORRE: é isto que diz qual descritor consertar ──
    atrator = Counter(deu for o in erros for _, deu, _, _ in erros[o])
    if atrator:
        print("\n── temas que ATRAEM erro (para onde o sistema manda o que não sabe) ──")
        for tema, n in atrator.most_common(5):
            print(f"   {n:>3}  {tema}")

    obs = [(sistema[d]['titulo'][:60], m['obs']) for d, m in gab.items() if m['obs']]
    if obs:
        print("\n── suas observações ──")
        for t, o in obs:
            print(f"   {t}\n      {o}")
    print("\n" + "═" * 76)
    print("O que este número decide: se o MeSH estiver muito abaixo do LLM, o conserto é no")
    print("mapa de descritores (src/dados/mesh_para_tema.json), não no prompt.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
